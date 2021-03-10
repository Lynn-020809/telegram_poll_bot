import json
import openpyxl
from shutil import copyfile
from openpyxl import load_workbook
from datetime import date

TEMPLATE_FP = "template.xlsx"
with open('club_members.json') as member_file:
    member_info = json.load(member_file)
with open('info.json') as info_file:
    poll_info = json.load(info_file)



def create_new_template(dst, src=TEMPLATE_FP):
    copyfile(src, dst)
    return load_workbook(filename=dst)


def excel_name():
    name_excel = "tt_"
    today = date.today()
    name_excel += today.strftime("%a %b-%d")
    name_excel += '.xlsx'
    return name_excel


def filling(day, wb):
    i = 3
    sheet = wb['All participants']
    for poll in poll_info:
        if int(day) in poll[1]:
            print('YES')
            j = 2
            user = poll[0]
            for information in member_info[user]:
                 sheet.cell(row = i, column = j).value = information
                 j += 1
            i += 1
        else:
            continue
    wb.save(excel_name())
        

def main():
    day = input("Key in the training number pls.")
    name_for_excel = excel_name()
    wb = create_new_template(name_for_excel)
    filling(day, wb)


main()
